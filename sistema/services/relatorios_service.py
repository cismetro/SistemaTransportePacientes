#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Transporte de Pacientes
Desenvolvido para a Prefeitura Municipal de Cosmópolis
Pode ser usado por terceiros autorizados

Serviços: Relatórios
Geração de relatórios em diversos formatos (PDF, Excel, CSV)
"""

from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional, Tuple
from io import BytesIO
import os
import tempfile
import json

# Bibliotecas para PDF
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor, black, white, grey
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.platypus.flowables import PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie

# Bibliotecas para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
import xlsxwriter

from flask import current_app
from sqlalchemy import and_, or_, func, desc

from sistema.models.agendamento import Agendamento
from sistema.models.paciente import Paciente
from sistema.models.motorista import Motorista
from sistema.models.veiculo import Veiculo
from sistema.models.usuario import Usuario
from sistema.status.status_enum import StatusAgendamento, TipoAtendimento, PrioridadeAgendamento
from db.database import db

class RelatoriosService:
    """
    Serviço para geração de relatórios
    Sistema da Prefeitura Municipal de Cosmópolis
    """
    
    def __init__(self):
        self.reports_dir = current_app.config.get('REPORTS_FOLDER', 'relatorios')
        self.colors = {
            'primary': HexColor('#4fc9c4'),
            'secondary': HexColor('#6d7a8c'),
            'success': HexColor('#28a745'),
            'warning': HexColor('#ffc107'),
            'danger': HexColor('#dc3545'),
            'info': HexColor('#17a2b8')
        }
        
        # Criar diretório se não existir
        os.makedirs(self.reports_dir, exist_ok=True)
    
    # === RELATÓRIOS GERAIS ===
    
    def gerar_relatorio_geral(self, formato: str = 'pdf', **filtros) -> Tuple[bool, str, Optional[bytes]]:
        """
        Gera relatório geral do sistema
        
        Args:
            formato: pdf, excel ou csv
            **filtros: Filtros para o relatório
            
        Returns:
            Tuple (sucesso, mensagem, conteúdo_bytes)
        """
        try:
            # Coletar dados
            dados = self._coletar_dados_gerais(**filtros)
            
            if formato.lower() == 'pdf':
                return self._gerar_pdf_geral(dados)
            elif formato.lower() == 'excel':
                return self._gerar_excel_geral(dados)
            elif formato.lower() == 'csv':
                return self._gerar_csv_geral(dados)
            else:
                return False, "Formato não suportado", None
                
        except Exception as e:
            current_app.logger.error(f"Erro ao gerar relatório geral: {e}")
            return False, f"Erro interno: {str(e)}", None
    
    def _coletar_dados_gerais(self, **filtros) -> Dict[str, Any]:
        """Coleta dados gerais do sistema"""
        
        data_inicio = filtros.get('data_inicio')
        data_fim = filtros.get('data_fim')
        
        # Se não especificado, usar último mês
        if not data_inicio:
            data_fim = date.today()
            data_inicio = data_fim - timedelta(days=30)
        elif isinstance(data_inicio, str):
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            if not data_fim:
                data_fim = date.today()
            elif isinstance(data_fim, str):
                data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        
        # Estatísticas gerais
        stats = {
            'pacientes': Paciente.estatisticas(),
            'motoristas': Motorista.estatisticas(),
            'veiculos': Veiculo.estatisticas(),
            'agendamentos': Agendamento.estatisticas()
        }
        
        # Agendamentos do período
        agendamentos_periodo = Agendamento.query.filter(
            and_(
                Agendamento.data_agendamento >= data_inicio,
                Agendamento.data_agendamento <= data_fim
            )
        ).all()
        
        # Análise por status
        por_status = {}
        for agendamento in agendamentos_periodo:
            status = agendamento.status
            por_status[status] = por_status.get(status, 0) + 1
        
        # Análise por tipo de atendimento
        por_tipo = {}
        for agendamento in agendamentos_periodo:
            tipo = agendamento.tipo_atendimento
            por_tipo[tipo] = por_tipo.get(tipo, 0) + 1
        
        # Análise por prioridade
        por_prioridade = {}
        for agendamento in agendamentos_periodo:
            prioridade = agendamento.prioridade
            por_prioridade[prioridade] = por_prioridade.get(prioridade, 0) + 1
        
        # Top motoristas (por número de viagens)
        top_motoristas = db.session.query(
            Motorista.nome_completo,
            func.count(Agendamento.id).label('total_viagens')
        ).join(Agendamento).filter(
            and_(
                Agendamento.data_agendamento >= data_inicio,
                Agendamento.data_agendamento <= data_fim,
                Agendamento.status == StatusAgendamento.CONCLUIDO.value
            )
        ).group_by(Motorista.id).order_by(desc('total_viagens')).limit(10).all()
        
        # Quilometragem total
        km_total = db.session.query(
            func.sum(Agendamento.km_percorrido)
        ).filter(
            and_(
                Agendamento.data_agendamento >= data_inicio,
                Agendamento.data_agendamento <= data_fim,
                Agendamento.km_percorrido != None
            )
        ).scalar() or 0
        
        return {
            'periodo': {
                'inicio': data_inicio,
                'fim': data_fim
            },
            'estatisticas_gerais': stats,
            'agendamentos_periodo': agendamentos_periodo,
            'por_status': por_status,
            'por_tipo': por_tipo,
            'por_prioridade': por_prioridade,
            'top_motoristas': top_motoristas,
            'km_total': km_total,
            'data_geracao': datetime.now()
        }
    
    # === GERAÇÃO PDF ===
    
    def _gerar_pdf_geral(self, dados: Dict[str, Any]) -> Tuple[bool, str, bytes]:
        """Gera relatório geral em PDF"""
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=self.colors['primary'],
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=self.colors['secondary'],
            spaceAfter=12
        )
        
        # Conteúdo
        story = []
        
        # Cabeçalho
        story.append(Paragraph("PREFEITURA MUNICIPAL DE COSMÓPOLIS", title_style))
        story.append(Paragraph("Sistema de Transporte de Pacientes", styles['Heading2']))
        story.append(Paragraph("Relatório Geral", styles['Heading2']))
        story.append(Spacer(1, 20))
        
        # Período
        periodo_texto = f"Período: {dados['periodo']['inicio'].strftime('%d/%m/%Y')} a {dados['periodo']['fim'].strftime('%d/%m/%Y')}"
        story.append(Paragraph(periodo_texto, styles['Normal']))
        story.append(Paragraph(f"Gerado em: {dados['data_geracao'].strftime('%d/%m/%Y às %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Estatísticas Gerais
        story.append(Paragraph("Estatísticas Gerais", heading_style))
        
        stats_data = [
            ['Categoria', 'Total', 'Ativos', 'Observações'],
            [
                'Pacientes',
                str(dados['estatisticas_gerais']['pacientes']['total']),
                str(dados['estatisticas_gerais']['pacientes']['ativos']),
                f"{dados['estatisticas_gerais']['pacientes']['cadeirantes']} cadeirantes"
            ],
            [
                'Motoristas',
                str(dados['estatisticas_gerais']['motoristas']['total']),
                str(dados['estatisticas_gerais']['motoristas']['ativos']),
                f"{dados['estatisticas_gerais']['motoristas']['habilitados']} habilitados"
            ],
            [
                'Veículos',
                str(dados['estatisticas_gerais']['veiculos']['total']),
                str(dados['estatisticas_gerais']['veiculos']['ativos']),
                f"{dados['estatisticas_gerais']['veiculos']['disponiveis']} disponíveis"
            ],
            [
                'Agendamentos',
                str(dados['estatisticas_gerais']['agendamentos']['total']),
                f"{dados['estatisticas_gerais']['agendamentos']['hoje_total']} hoje",
                f"{dados['estatisticas_gerais']['agendamentos']['concluidos']} concluídos"
            ]
        ]
        
        stats_table = Table(stats_data, colWidths=[4*cm, 2*cm, 2*cm, 4*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.colors['primary']),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), white),
            ('GRID', (0, 0), (-1, -1), 1, black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # Agendamentos por Status
        if dados['por_status']:
            story.append(Paragraph("Agendamentos por Status", heading_style))
            
            status_data = [['Status', 'Quantidade', 'Percentual']]
            total_agendamentos = len(dados['agendamentos_periodo'])
            
            for status, quantidade in dados['por_status'].items():
                percentual = (quantidade / total_agendamentos * 100) if total_agendamentos > 0 else 0
                status_nome = StatusAgendamento.get_display_name(status)
                status_data.append([status_nome, str(quantidade), f"{percentual:.1f}%"])
            
            status_table = Table(status_data, colWidths=[6*cm, 3*cm, 3*cm])
            status_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors['info']),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, black)
            ]))
            
            story.append(status_table)
            story.append(Spacer(1, 20))
        
        # Top Motoristas
        if dados['top_motoristas']:
            story.append(Paragraph("Top 10 Motoristas (Viagens Concluídas)", heading_style))
            
            motoristas_data = [['Posição', 'Motorista', 'Viagens']]
            for i, (nome, viagens) in enumerate(dados['top_motoristas'], 1):
                motoristas_data.append([str(i), nome, str(viagens)])
            
            motoristas_table = Table(motoristas_data, colWidths=[2*cm, 8*cm, 2*cm])
            motoristas_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors['success']),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, black)
            ]))
            
            story.append(motoristas_table)
            story.append(Spacer(1, 20))
        
        # Quilometragem
        story.append(Paragraph("Resumo de Quilometragem", heading_style))
        story.append(Paragraph(f"Total de quilômetros percorridos no período: {dados['km_total']:.0f} km", styles['Normal']))
        story.append(Spacer(1, 10))
        
        # Rodapé
        story.append(PageBreak())
        story.append(Paragraph("_______________________________________________", styles['Normal']))
        story.append(Paragraph("Prefeitura Municipal de Cosmópolis", styles['Normal']))
        story.append(Paragraph("Sistema de Transporte de Pacientes", styles['Normal']))
        
        # Gerar PDF
        doc.build(story)
        buffer.seek(0)
        
        return True, "Relatório PDF gerado com sucesso", buffer.getvalue()
    
    # === GERAÇÃO EXCEL ===
    
    def _gerar_excel_geral(self, dados: Dict[str, Any]) -> Tuple[bool, str, bytes]:
        """Gera relatório geral em Excel"""
        
        buffer = BytesIO()
        workbook = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4FC9C4", end_color="4FC9C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aba 1: Resumo
        ws_resumo = workbook.active
        ws_resumo.title = "Resumo"
        
        # Cabeçalho
        ws_resumo['A1'] = "PREFEITURA MUNICIPAL DE COSMÓPOLIS"
        ws_resumo['A2'] = "Sistema de Transporte de Pacientes"
        ws_resumo['A3'] = "Relatório Geral"
        
        ws_resumo['A1'].font = Font(bold=True, size=16)
        ws_resumo['A2'].font = Font(bold=True, size=12)
        ws_resumo['A3'].font = Font(bold=True, size=12)
        
        # Período
        ws_resumo['A5'] = "Período:"
        ws_resumo['B5'] = f"{dados['periodo']['inicio'].strftime('%d/%m/%Y')} a {dados['periodo']['fim'].strftime('%d/%m/%Y')}"
        ws_resumo['A6'] = "Gerado em:"
        ws_resumo['B6'] = dados['data_geracao'].strftime('%d/%m/%Y às %H:%M')
        
        # Estatísticas Gerais
        row = 8
        ws_resumo[f'A{row}'] = "Estatísticas Gerais"
        ws_resumo[f'A{row}'].font = header_font
        ws_resumo[f'A{row}'].fill = header_fill
        
        row += 1
        headers = ['Categoria', 'Total', 'Ativos', 'Observações']
        for col, header in enumerate(headers, 1):
            cell = ws_resumo.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados das estatísticas
        stats_data = [
            ['Pacientes', dados['estatisticas_gerais']['pacientes']['total'], 
             dados['estatisticas_gerais']['pacientes']['ativos'], 
             f"{dados['estatisticas_gerais']['pacientes']['cadeirantes']} cadeirantes"],
            ['Motoristas', dados['estatisticas_gerais']['motoristas']['total'], 
             dados['estatisticas_gerais']['motoristas']['ativos'], 
             f"{dados['estatisticas_gerais']['motoristas']['habilitados']} habilitados"],
            ['Veículos', dados['estatisticas_gerais']['veiculos']['total'], 
             dados['estatisticas_gerais']['veiculos']['ativos'], 
             f"{dados['estatisticas_gerais']['veiculos']['disponiveis']} disponíveis"],
            ['Agendamentos', dados['estatisticas_gerais']['agendamentos']['total'], 
             f"{dados['estatisticas_gerais']['agendamentos']['hoje_total']} hoje", 
             f"{dados['estatisticas_gerais']['agendamentos']['concluidos']} concluídos"]
        ]
        
        for data_row in stats_data:
            row += 1
            for col, value in enumerate(data_row, 1):
                cell = ws_resumo.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Aba 2: Agendamentos por Status
        if dados['por_status']:
            ws_status = workbook.create_sheet("Por Status")
            
            ws_status['A1'] = "Agendamentos por Status"
            ws_status['A1'].font = Font(bold=True, size=14)
            
            # Headers
            ws_status['A3'] = "Status"
            ws_status['B3'] = "Quantidade"
            ws_status['C3'] = "Percentual"
            
            for cell in ['A3', 'B3', 'C3']:
                ws_status[cell].font = header_font
                ws_status[cell].fill = header_fill
                ws_status[cell].border = border
            
            # Dados
            row = 4
            total_agendamentos = len(dados['agendamentos_periodo'])
            
            for status, quantidade in dados['por_status'].items():
                percentual = (quantidade / total_agendamentos * 100) if total_agendamentos > 0 else 0
                status_nome = StatusAgendamento.get_display_name(status)
                
                ws_status[f'A{row}'] = status_nome
                ws_status[f'B{row}'] = quantidade
                ws_status[f'C{row}'] = f"{percentual:.1f}%"
                
                for col in ['A', 'B', 'C']:
                    ws_status[f'{col}{row}'].border = border
                
                row += 1
        
        # Aba 3: Top Motoristas
        if dados['top_motoristas']:
            ws_motoristas = workbook.create_sheet("Top Motoristas")
            
            ws_motoristas['A1'] = "Top 10 Motoristas (Viagens Concluídas)"
            ws_motoristas['A1'].font = Font(bold=True, size=14)
            
            # Headers
            ws_motoristas['A3'] = "Posição"
            ws_motoristas['B3'] = "Motorista"
            ws_motoristas['C3'] = "Viagens"
            
            for cell in ['A3', 'B3', 'C3']:
                ws_motoristas[cell].font = header_font
                ws_motoristas[cell].fill = header_fill
                ws_motoristas[cell].border = border
            
            # Dados
            row = 4
            for i, (nome, viagens) in enumerate(dados['top_motoristas'], 1):
                ws_motoristas[f'A{row}'] = i
                ws_motoristas[f'B{row}'] = nome
                ws_motoristas[f'C{row}'] = viagens
                
                for col in ['A', 'B', 'C']:
                    ws_motoristas[f'{col}{row}'].border = border
                
                row += 1
        
        # Ajustar larguras das colunas
        for ws in workbook.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar
        workbook.save(buffer)
        buffer.seek(0)
        
        return True, "Relatório Excel gerado com sucesso", buffer.getvalue()
    
    # === RELATÓRIOS ESPECÍFICOS ===
    
    def gerar_relatorio_agendamentos(self, formato: str = 'pdf', **filtros) -> Tuple[bool, str, bytes]:
        """Gera relatório específico de agendamentos"""
        
        try:
            # Aplicar filtros
            query = Agendamento.query
            
            if filtros.get('data_inicio'):
                data_inicio = datetime.strptime(filtros['data_inicio'], '%Y-%m-%d').date()
                query = query.filter(Agendamento.data_agendamento >= data_inicio)
            
            if filtros.get('data_fim'):
                data_fim = datetime.strptime(filtros['data_fim'], '%Y-%m-%d').date()
                query = query.filter(Agendamento.data_agendamento <= data_fim)
            
            if filtros.get('status'):
                query = query.filter(Agendamento.status == filtros['status'])
            
            if filtros.get('motorista_id'):
                query = query.filter(Agendamento.motorista_id == filtros['motorista_id'])
            
            if filtros.get('veiculo_id'):
                query = query.filter(Agendamento.veiculo_id == filtros['veiculo_id'])
            
            agendamentos = query.order_by(
                Agendamento.data_agendamento,
                Agendamento.horario_saida
            ).all()
            
            if formato.lower() == 'pdf':
                return self._gerar_pdf_agendamentos(agendamentos, filtros)
            elif formato.lower() == 'excel':
                return self._gerar_excel_agendamentos(agendamentos, filtros)
            else:
                return False, "Formato não suportado", None
                
        except Exception as e:
            current_app.logger.error(f"Erro ao gerar relatório de agendamentos: {e}")
            return False, f"Erro interno: {str(e)}", None
    
    def _gerar_pdf_agendamentos(self, agendamentos: List[Agendamento], filtros: Dict) -> Tuple[bool, str, bytes]:
        """Gera relatório de agendamentos em PDF"""
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Cabeçalho
        story.append(Paragraph("PREFEITURA MUNICIPAL DE COSMÓPOLIS", styles['Title']))
        story.append(Paragraph("Relatório de Agendamentos", styles['Heading1']))
        story.append(Spacer(1, 20))
        
        # Filtros aplicados
        if filtros:
            story.append(Paragraph("Filtros aplicados:", styles['Heading2']))
            for chave, valor in filtros.items():
                if valor:
                    story.append(Paragraph(f"• {chave.replace('_', ' ').title()}: {valor}", styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Dados dos agendamentos
        if agendamentos:
            data = [['Data', 'Horário', 'Paciente', 'Destino', 'Status']]
            
            for agendamento in agendamentos:
                data.append([
                    agendamento.data_agendamento.strftime('%d/%m/%Y'),
                    agendamento.horario_saida.strftime('%H:%M'),
                    agendamento.paciente.nome_completo[:30],
                    agendamento.destino_nome[:30],
                    StatusAgendamento.get_display_name(agendamento.status)
                ])
            
            table = Table(data, colWidths=[2*cm, 2*cm, 4*cm, 4*cm, 2*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors['primary']),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, black)
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("Nenhum agendamento encontrado com os filtros aplicados.", styles['Normal']))
        
        # Resumo
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Total de agendamentos: {len(agendamentos)}", styles['Normal']))
        story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        
        return True, "Relatório de agendamentos gerado com sucesso", buffer.getvalue()
    
    def _gerar_excel_agendamentos(self, agendamentos: List[Agendamento], filtros: Dict) -> Tuple[bool, str, bytes]:
        """Gera relatório de agendamentos em Excel"""
        
        buffer = BytesIO()
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Agendamentos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4FC9C4", end_color="4FC9C4", fill_type="solid")
        
        # Cabeçalho
        ws['A1'] = "Relatório de Agendamentos"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Headers da tabela
        headers = [
            'Data', 'Horário Saída', 'Horário Retorno', 'Paciente', 'CPF Paciente',
            'Motorista', 'Veículo', 'Destino', 'Tipo Atendimento', 'Prioridade', 'Status'
        ]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Dados
        for agendamento in agendamentos:
            row += 1
            dados_linha = [
                agendamento.data_agendamento.strftime('%d/%m/%Y'),
                agendamento.horario_saida.strftime('%H:%M'),
                agendamento.horario_retorno_previsto.strftime('%H:%M') if agendamento.horario_retorno_previsto else '',
                agendamento.paciente.nome_completo,
                agendamento.paciente.cpf_formatado,
                agendamento.motorista.nome_completo,
                agendamento.veiculo.placa_formatada,
                agendamento.destino_nome,
                TipoAtendimento.get_display_name(agendamento.tipo_atendimento),
                PrioridadeAgendamento.get_display_name(agendamento.prioridade),
                StatusAgendamento.get_display_name(agendamento.status)
            ]
            
            for col, valor in enumerate(dados_linha, 1):
                ws.cell(row=row, column=col, value=valor)
        
        # Ajustar larguras
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(buffer)
        buffer.seek(0)
        
        return True, "Relatório Excel de agendamentos gerado com sucesso", buffer.getvalue()
    
    # === RELATÓRIOS DE CONTROLE ===
    
    def gerar_relatorio_controle_vencimentos(self, formato: str = 'pdf') -> Tuple[bool, str, bytes]:
        """Gera relatório de controle de vencimentos (CNH, licenciamento, etc.)"""
        
        try:
            # Coletar dados de vencimentos
            dados_vencimentos = self._coletar_dados_vencimentos()
            
            if formato.lower() == 'pdf':
                return self._gerar_pdf_vencimentos(dados_vencimentos)
            elif formato.lower() == 'excel':
                return self._gerar_excel_vencimentos(dados_vencimentos)
            else:
                return False, "Formato não suportado", None
                
        except Exception as e:
            current_app.logger.error(f"Erro ao gerar relatório de vencimentos: {e}")
            return False, f"Erro interno: {str(e)}", None
    
    def _coletar_dados_vencimentos(self) -> Dict[str, Any]:
        """Coleta dados de vencimentos"""
        
        hoje = date.today()
        limite_30_dias = hoje + timedelta(days=30)
        
        # CNH vencendo/vencida
        cnh_vencendo = Motorista.query.filter(
            and_(
                Motorista.ativo == True,
                Motorista.data_vencimento_cnh <= limite_30_dias,
                Motorista.data_vencimento_cnh >= hoje
            )
        ).all()
        
        cnh_vencida = Motorista.query.filter(
            and_(
                Motorista.ativo == True,
                Motorista.data_vencimento_cnh < hoje
            )
        ).all()
        
        # Licenciamento vencendo/vencido
        licenciamento_vencendo = Veiculo.query.filter(
            and_(
                Veiculo.ativo == True,
                Veiculo.data_vencimento_licenciamento <= limite_30_dias,
                Veiculo.data_vencimento_licenciamento >= hoje
            )
        ).all()
        
        licenciamento_vencido = Veiculo.query.filter(
            and_(
                Veiculo.ativo == True,
                Veiculo.data_vencimento_licenciamento < hoje
            )
        ).all()
        
        # Veículos que precisam de revisão
        necessitam_revisao = Veiculo.necessitam_revisao().all()
        
        return {
            'cnh_vencendo': cnh_vencendo,
            'cnh_vencida': cnh_vencida,
            'licenciamento_vencendo': licenciamento_vencendo,
            'licenciamento_vencido': licenciamento_vencido,
            'necessitam_revisao': necessitam_revisao,
            'data_geracao': datetime.now()
        }
    
    def _gerar_pdf_vencimentos(self, dados: Dict[str, Any]) -> Tuple[bool, str, bytes]:
        """Gera relatório de vencimentos em PDF"""
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Cabeçalho
        story.append(Paragraph("PREFEITURA MUNICIPAL DE COSMÓPOLIS", styles['Title']))
        story.append(Paragraph("Relatório de Controle de Vencimentos", styles['Heading1']))
        story.append(Spacer(1, 20))
        
        # CNH Vencida (URGENTE)
        if dados['cnh_vencida']:
            story.append(Paragraph("🚨 CNH VENCIDA - AÇÃO URGENTE", styles['Heading2']))
            
            cnh_data = [['Motorista', 'CNH', 'Vencimento', 'Dias Vencida']]
            for motorista in dados['cnh_vencida']:
                dias_vencida = (date.today() - motorista.data_vencimento_cnh).days
                cnh_data.append([
                    motorista.nome_completo,
                    motorista.numero_cnh,
                    motorista.data_vencimento_cnh.strftime('%d/%m/%Y'),
                    str(dias_vencida)
                ])
            
            table = Table(cnh_data, colWidths=[5*cm, 3*cm, 2.5*cm, 2.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors['danger']),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 15))
        
        # CNH Vencendo
        if dados['cnh_vencendo']:
            story.append(Paragraph("⚠️ CNH VENCENDO (próximos 30 dias)", styles['Heading2']))
            
            cnh_data = [['Motorista', 'CNH', 'Vencimento', 'Dias Restantes']]
            for motorista in dados['cnh_vencendo']:
                dias_restantes = (motorista.data_vencimento_cnh - date.today()).days
                cnh_data.append([
                    motorista.nome_completo,
                    motorista.numero_cnh,
                    motorista.data_vencimento_cnh.strftime('%d/%m/%Y'),
                    str(dias_restantes)
                ])
            
            table = Table(cnh_data, colWidths=[5*cm, 3*cm, 2.5*cm, 2.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors['warning']),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 15))
        
        # Licenciamento Vencido
        if dados['licenciamento_vencido']:
            story.append(Paragraph("🚨 LICENCIAMENTO VENCIDO", styles['Heading2']))
            
            lic_data = [['Veículo', 'Placa', 'Vencimento', 'Dias Vencido']]
            for veiculo in dados['licenciamento_vencido']:
                dias_vencido = (date.today() - veiculo.data_vencimento_licenciamento).days
                lic_data.append([
                    f"{veiculo.marca} {veiculo.modelo}",
                    veiculo.placa_formatada,
                    veiculo.data_vencimento_licenciamento.strftime('%d/%m/%Y'),
                    str(dias_vencido)
                ])
            
            table = Table(lic_data, colWidths=[5*cm, 2*cm, 2.5*cm, 2.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors['danger']),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 15))
        
        # Licenciamento Vencendo
        if dados['licenciamento_vencendo']:
            story.append(Paragraph("⚠️ LICENCIAMENTO VENCENDO", styles['Heading2']))
            
            lic_data = [['Veículo', 'Placa', 'Vencimento', 'Dias Restantes']]
            for veiculo in dados['licenciamento_vencendo']:
                dias_restantes = (veiculo.data_vencimento_licenciamento - date.today()).days
                lic_data.append([
                    f"{veiculo.marca} {veiculo.modelo}",
                    veiculo.placa_formatada,
                    veiculo.data_vencimento_licenciamento.strftime('%d/%m/%Y'),
                    str(dias_restantes)
                ])
            
            table = Table(lic_data, colWidths=[5*cm, 2*cm, 2.5*cm, 2.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors['warning']),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 15))
        
        # Veículos que precisam de revisão
        if dados['necessitam_revisao']:
            story.append(Paragraph("🔧 VEÍCULOS QUE PRECISAM DE REVISÃO", styles['Heading2']))
            
            rev_data = [['Veículo', 'Placa', 'KM Atual', 'Última Revisão']]
            for veiculo in dados['necessitam_revisao']:
                ultima_revisao = veiculo.data_ultima_revisao.strftime('%d/%m/%Y') if veiculo.data_ultima_revisao else 'N/A'
                rev_data.append([
                    f"{veiculo.marca} {veiculo.modelo}",
                    veiculo.placa_formatada,
                    str(veiculo.quilometragem_atual),
                    ultima_revisao
                ])
            
            table = Table(rev_data, colWidths=[5*cm, 2*cm, 2*cm, 3*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors['info']),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, black)
            ]))
            
            story.append(table)
        
        # Rodapé
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Gerado em: {dados['data_geracao'].strftime('%d/%m/%Y às %H:%M')}", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        
        return True, "Relatório de vencimentos gerado com sucesso", buffer.getvalue()
    
    # === GERAÇÃO CSV ===
    
    def _gerar_csv_geral(self, dados: Dict[str, Any]) -> Tuple[bool, str, bytes]:
        """Gera relatório geral em CSV"""
        
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Cabeçalho
        writer.writerow(['PREFEITURA MUNICIPAL DE COSMÓPOLIS'])
        writer.writerow(['Relatório Geral - Sistema de Transporte de Pacientes'])
        writer.writerow([])
        writer.writerow([f"Período: {dados['periodo']['inicio'].strftime('%d/%m/%Y')} a {dados['periodo']['fim'].strftime('%d/%m/%Y')}"])
        writer.writerow([f"Gerado em: {dados['data_geracao'].strftime('%d/%m/%Y às %H:%M')}"])
        writer.writerow([])
        
        # Estatísticas Gerais
        writer.writerow(['ESTATÍSTICAS GERAIS'])
        writer.writerow(['Categoria', 'Total', 'Ativos', 'Observações'])
        
        stats = dados['estatisticas_gerais']
        writer.writerow([
            'Pacientes', stats['pacientes']['total'], stats['pacientes']['ativos'],
            f"{stats['pacientes']['cadeirantes']} cadeirantes"
        ])
        writer.writerow([
            'Motoristas', stats['motoristas']['total'], stats['motoristas']['ativos'],
            f"{stats['motoristas']['habilitados']} habilitados"
        ])
        writer.writerow([
            'Veículos', stats['veiculos']['total'], stats['veiculos']['ativos'],
            f"{stats['veiculos']['disponiveis']} disponíveis"
        ])
        writer.writerow([
            'Agendamentos', stats['agendamentos']['total'], 
            f"{stats['agendamentos']['hoje_total']} hoje",
            f"{stats['agendamentos']['concluidos']} concluídos"
        ])
        
        writer.writerow([])
        
        # Agendamentos por Status
        if dados['por_status']:
            writer.writerow(['AGENDAMENTOS POR STATUS'])
            writer.writerow(['Status', 'Quantidade', 'Percentual'])
            
            total_agendamentos = len(dados['agendamentos_periodo'])
            for status, quantidade in dados['por_status'].items():
                percentual = (quantidade / total_agendamentos * 100) if total_agendamentos > 0 else 0
                status_nome = StatusAgendamento.get_display_name(status)
                writer.writerow([status_nome, quantidade, f"{percentual:.1f}%"])
        
        writer.writerow([])
        
        # Top Motoristas
        if dados['top_motoristas']:
            writer.writerow(['TOP 10 MOTORISTAS (Viagens Concluídas)'])
            writer.writerow(['Posição', 'Motorista', 'Viagens'])
            
            for i, (nome, viagens) in enumerate(dados['top_motoristas'], 1):
                writer.writerow([i, nome, viagens])
        
        writer.writerow([])
        writer.writerow(['QUILOMETRAGEM'])
        writer.writerow([f"Total de quilômetros percorridos: {dados['km_total']:.0f} km"])
        
        # Converter para bytes
        output.seek(0)
        csv_content = output.getvalue().encode('utf-8-sig')  # BOM para Excel
        
        return True, "Relatório CSV gerado com sucesso", csv_content
    
    # === UTILITÁRIOS ===
    
    def salvar_relatorio(self, conteudo: bytes, nome_arquivo: str) -> str:
        """Salva relatório no diretório configurado"""
        
        caminho_completo = os.path.join(self.reports_dir, nome_arquivo)
        
        with open(caminho_completo, 'wb') as f:
            f.write(conteudo)
        
        return caminho_completo
    
    def limpar_relatorios_antigos(self, dias: int = 30):
        """Remove relatórios mais antigos que X dias"""
        
        if not os.path.exists(self.reports_dir):
            return
        
        data_limite = datetime.now() - timedelta(days=dias)
        
        for arquivo in os.listdir(self.reports_dir):
            caminho = os.path.join(self.reports_dir, arquivo)
            
            if os.path.isfile(caminho):
                data_criacao = datetime.fromtimestamp(os.path.getctime(caminho))
                
                if data_criacao < data_limite:
                    try:
                        os.remove(caminho)
                        current_app.logger.info(f"Relatório antigo removido: {arquivo}")
                    except Exception as e:
                        current_app.logger.error(f"Erro ao remover relatório {arquivo}: {e}")

# Instância global do serviço
relatorios_service = RelatoriosService()